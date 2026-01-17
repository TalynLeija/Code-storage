#https://www.youtube.com/watch?v=aBgibCSXSoQ&t=1s
import random
import openpyxl
wbook = openpyxl.Workbook()
# import Robs_Coffe_Rules as RC

#you only need to execute this program once
wbook = openpyxl.Workbook()
ws = wbook.active
# ws.create_sheet('addition_data.xlsx')
#method
def sum(list):
    sum=0
    for i in list:
        sum+=i
    return sum

#labels
Heading = ['sum','num1','num2','num3','num4','num5']
ws.append(Heading)

#generateing data
random.seed(5)
data = []
for i in range(200):
    templist=[]
    for j in range(5):
        templist.append(random.randint(1, 21))
    data.append(templist)
# print(data)
#add to excel
for i in range(len(data)):
    ws.append([sum(data[i]),data[i][0],data[i][1],data[i][2],data[i][3],data[i][4]])

wbook.save('addition_data.xlsx') 
print("All is well")